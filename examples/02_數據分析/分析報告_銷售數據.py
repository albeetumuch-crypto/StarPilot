from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import seaborn as sns
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

# 設定中文字體
plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
sns.set_style("whitegrid")

# 讀取數據
csv_path = './測試資料/sales_data.csv'
df = pd.read_csv(csv_path)

# 轉換日期格式
df['日期'] = pd.to_datetime(df['日期'])
df['月份'] = df['日期'].dt.to_period('M')
df['年月'] = df['日期'].dt.strftime('%Y-%m')

print("📊 開始分析銷售數據...")
print(f"✅ 讀取 {len(df)} 筆銷售記錄")

# ==================== 數據統計 ====================
total_sales = df['金額'].sum()
avg_sales = df['金額'].mean()
total_orders = len(df)
num_products = df['產品'].nunique()
num_regions = df['地區'].nunique()
num_salespeople = df['業務'].nunique()

print(f"\n📈 基本統計：")
print(f"   總銷售額: ${total_sales:,.0f}")
print(f"   平均訂單金額: ${avg_sales:,.0f}")
print(f"   訂單總數: {total_orders}")
print(f"   產品種類: {num_products}")
print(f"   銷售地區: {num_regions}")
print(f"   業務人員: {num_salespeople}")

# ==================== 各維度分析 ====================
# 產品分析
product_stats = df.groupby('產品').agg({
    '金額': ['sum', 'mean', 'count'],
    '數量': 'sum'
}).round(2)
product_stats.columns = ['總金額', '平均金額', '訂單數', '總數量']
product_stats = product_stats.sort_values('總金額', ascending=False)

# 地區分析
region_stats = df.groupby('地區').agg({
    '金額': ['sum', 'mean', 'count'],
    '數量': 'sum'
}).round(2)
region_stats.columns = ['總金額', '平均金額', '訂單數', '總數量']
region_stats = region_stats.sort_values('總金額', ascending=False)

# 業務分析
sales_stats = df.groupby('業務').agg({
    '金額': ['sum', 'mean', 'count'],
    '數量': 'sum'
}).round(2)
sales_stats.columns = ['總金額', '平均金額', '訂單數', '總數量']
sales_stats = sales_stats.sort_values('總金額', ascending=False)

# 月份分析
monthly_stats = df.groupby('年月').agg({
    '金額': ['sum', 'mean', 'count'],
    '數量': 'sum'
}).round(2)
monthly_stats.columns = ['總金額', '平均金額', '訂單數', '總數量']

print(f"\n🏆 產品排名 (前 5)：")
print(product_stats.head())

# ==================== 生成圖表 ====================
print("\n🎨 生成圖表中...")

# 清空之前的圖表
plt.close('all')

# 圖表1: 產品銷售額
fig1, ax1 = plt.subplots(figsize=(10, 6))
product_sales = df.groupby('產品')['金額'].sum().sort_values(ascending=False)
colors = sns.color_palette("husl", len(product_sales))
ax1.barh(product_sales.index, product_sales.values, color=colors)
ax1.set_xlabel('銷售額 ($)', fontsize=12)
ax1.set_title('各產品銷售總額', fontsize=14, fontweight='bold')
ax1.invert_yaxis()
for i, v in enumerate(product_sales.values):
    ax1.text(v, i, f' ${v:,.0f}', va='center', fontsize=10)
plt.tight_layout()
chart1_bytes = BytesIO()
fig1.savefig(chart1_bytes, format='png', dpi=100, bbox_inches='tight')
chart1_bytes.seek(0)

# 圖表2: 地區銷售分布
fig2, ax2 = plt.subplots(figsize=(10, 6))
region_sales = df.groupby('地區')['金額'].sum().sort_values(ascending=False)
colors = sns.color_palette("Set2", len(region_sales))
wedges, texts, autotexts = ax2.pie(region_sales.values, labels=region_sales.index,
                                   autopct='%1.1f%%', colors=colors, startangle=90)
ax2.set_title('各地區銷售分布', fontsize=14, fontweight='bold')
for autotext in autotexts:
    autotext.set_color('white')
    autotext.set_fontsize(11)
    autotext.set_fontweight('bold')
plt.tight_layout()
chart2_bytes = BytesIO()
fig2.savefig(chart2_bytes, format='png', dpi=100, bbox_inches='tight')
chart2_bytes.seek(0)

# 圖表3: 月份銷售趨勢
fig3, ax3 = plt.subplots(figsize=(12, 6))
monthly_sales = df.groupby('年月')['金額'].sum()
ax3.plot(range(len(monthly_sales)), monthly_sales.values,
         marker='o', linewidth=2.5, markersize=8, color='steelblue')
ax3.fill_between(range(len(monthly_sales)),
                 monthly_sales.values, alpha=0.3, color='steelblue')
ax3.set_xticks(range(len(monthly_sales)))
ax3.set_xticklabels(monthly_sales.index, rotation=45)
ax3.set_ylabel('銷售額 ($)', fontsize=12)
ax3.set_title('月度銷售趨勢', fontsize=14, fontweight='bold')
ax3.grid(True, alpha=0.3)
for i, v in enumerate(monthly_sales.values):
    ax3.text(i, v, f'${v:,.0f}', ha='center', va='bottom', fontsize=10)
plt.tight_layout()
chart3_bytes = BytesIO()
fig3.savefig(chart3_bytes, format='png', dpi=100, bbox_inches='tight')
chart3_bytes.seek(0)

# 圖表4: 業務銷售額
fig4, ax4 = plt.subplots(figsize=(10, 6))
sales_personnel = df.groupby('業務')['金額'].sum().sort_values(ascending=False)
colors = sns.color_palette("RdYlGn", len(sales_personnel))
bars = ax4.bar(sales_personnel.index, sales_personnel.values, color=colors)
ax4.set_ylabel('銷售額 ($)', fontsize=12)
ax4.set_title('各業務人員銷售總額', fontsize=14, fontweight='bold')
ax4.tick_params(axis='x', rotation=45)
for bar in bars:
    height = bar.get_height()
    ax4.text(bar.get_x() + bar.get_width()/2., height,
             f'${height:,.0f}',
             ha='center', va='bottom', fontsize=10)
plt.tight_layout()
chart4_bytes = BytesIO()
fig4.savefig(chart4_bytes, format='png', dpi=100, bbox_inches='tight')
chart4_bytes.seek(0)

# ==================== 生成 Excel 文件 ====================
print("\n📁 生成 Excel 報告...")


# 創建工作簿
wb = Workbook()
wb.remove(wb.active)  # 删除默认工作表

# ===== 工作表1: 摘要頁 =====
ws_summary = wb.create_sheet("摘要", 0)
ws_summary.sheet_properties.tabColor = "FFFF00"

# 標題
ws_summary['A1'] = '銷售數據分析報告'
ws_summary['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_summary['A1'].fill = PatternFill(
    start_color="366092", end_color="366092", fill_type="solid")
ws_summary.merge_cells('A1:D1')
ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_summary.row_dimensions[1].height = 30

# 報告日期
ws_summary['A2'] = f"報告日期：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws_summary['A2'].font = Font(size=11, italic=True)

# 統計數據
row = 4
headers = ['指標', '數值']
for col, header in enumerate(headers, 1):
    cell = ws_summary.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="4472C4",
                            end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

data = [
    ['總銷售額', f'${total_sales:,.0f}'],
    ['平均訂單金額', f'${avg_sales:,.0f}'],
    ['訂單總數', f'{total_orders}'],
    ['產品種類', f'{num_products}'],
    ['銷售地區', f'{num_regions}'],
    ['業務人員', f'{num_salespeople}'],
]

for idx, row_data in enumerate(data, 5):
    ws_summary.cell(row=idx, column=1,
                    value=row_data[0]).font = Font(bold=True)
    ws_summary.cell(row=idx, column=2, value=row_data[1]).font = Font(size=11)
    if idx % 2 == 0:
        for col in range(1, 3):
            ws_summary.cell(row=idx, column=col).fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

ws_summary.column_dimensions['A'].width = 15
ws_summary.column_dimensions['B'].width = 20

# 插入圖表1
img1 = XLImage(chart1_bytes)
img1.width = 500
img1.height = 300
ws_summary.add_image(img1, 'A12')

# ===== 工作表2: 原始數據 =====
ws_data = wb.create_sheet("原始數據", 1)
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        else:
            if r_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col in ws_data.columns:
    max_length = 0
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws_data.column_dimensions[get_column_letter(
        col[0].column)].width = adjusted_width

# ===== 工作表3: 產品分析 =====
ws_product = wb.create_sheet("產品分析", 2)
ws_product['A1'] = '產品分析統計'
ws_product['A1'].font = Font(size=14, bold=True)
ws_product.merge_cells('A1:E1')

for r_idx, row in enumerate(dataframe_to_rows(product_stats.reset_index(), index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_product.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            if r_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col in range(1, 6):
    ws_product.column_dimensions[get_column_letter(col)].width = 15

# 插入產品圖表
img_product = XLImage(chart1_bytes)
img_product.width = 500
img_product.height = 300
ws_product.add_image(img_product, 'A10')

# ===== 工作表4: 地區分析 =====
ws_region = wb.create_sheet("地區分析", 3)
ws_region['A1'] = '地區分析統計'
ws_region['A1'].font = Font(size=14, bold=True)
ws_region.merge_cells('A1:E1')

for r_idx, row in enumerate(dataframe_to_rows(region_stats.reset_index(), index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_region.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            if r_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col in range(1, 6):
    ws_region.column_dimensions[get_column_letter(col)].width = 15

# 插入圖表2
img2 = XLImage(chart2_bytes)
img2.width = 400
img2.height = 400
ws_region.add_image(img2, 'A10')

# ===== 工作表5: 月份趨勢 =====
ws_monthly = wb.create_sheet("月份趨勢", 4)
ws_monthly['A1'] = '月度銷售趨勢'
ws_monthly['A1'].font = Font(size=14, bold=True)
ws_monthly.merge_cells('A1:E1')

for r_idx, row in enumerate(dataframe_to_rows(monthly_stats.reset_index(), index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_monthly.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            if r_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col in range(1, 6):
    ws_monthly.column_dimensions[get_column_letter(col)].width = 15

# 插入圖表3
img3 = XLImage(chart3_bytes)
img3.width = 600
img3.height = 300
ws_monthly.add_image(img3, 'A10')

# ===== 工作表6: 業務分析 =====
ws_sales = wb.create_sheet("業務分析", 5)
ws_sales['A1'] = '業務人員分析統計'
ws_sales['A1'].font = Font(size=14, bold=True)
ws_sales.merge_cells('A1:E1')

for r_idx, row in enumerate(dataframe_to_rows(sales_stats.reset_index(), index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_sales.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            if r_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col in range(1, 6):
    ws_sales.column_dimensions[get_column_letter(col)].width = 15

# 插入圖表4
img4 = XLImage(chart4_bytes)
img4.width = 500
img4.height = 300
ws_sales.add_image(img4, 'A10')

# ===== 保存文件 =====
output_path = './銷售數據分析報告.xlsx'
wb.save(output_path)

print(f"\n✅ Excel 報告已生成：{output_path}")
print(f"\n📊 報告包含以下工作表：")
print(f"   1️⃣  摘要 - 關鍵統計數據和產品銷售排名")
print(f"   2️⃣  原始數據 - 所有銷售交易記錄")
print(f"   3️⃣  產品分析 - 各產品銷售統計與排行")
print(f"   4️⃣  地區分析 - 各地區銷售分布")
print(f"   5️⃣  月份趨勢 - 月度銷售走勢分析")
print(f"   6️⃣  業務分析 - 業務人員績效統計")
print("\n✨ 分析完成！")
